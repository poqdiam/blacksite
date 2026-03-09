"""
CIS Controls v8 IG Baseline Import
====================================
Reads the official CIS Controls v8 Excel spreadsheet and populates:
  - catalog_controls  — descriptions + metadata for all 153 CIS v8 safeguards
                        (matches existing rows by control_id "CIS X.Y")
  - baseline_controls — IG1 / IG2 / IG3 membership (cumulative: IG1 ⊂ IG2 ⊂ IG3)

Also fixes cis8 framework kind: 'framework' → 'catalog'

Usage:
    python scripts/import_cis_ig.py [path/to/CIS_Controls_Version_8.xlsx]

Defaults to controls/CIS_Controls_Version_8.xlsx if no path given.
Safe to re-run — all inserts use INSERT OR IGNORE.
"""

from __future__ import annotations
import sys
import sqlite3
import logging
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)

ROOT    = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "blacksite.db"

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger("import_cis_ig")

SHEET_NAME = "Controls V8"
DEFAULT_XLSX = ROOT / "controls" / "CIS_Controls_Version_8.xlsx"

# Column indices (0-based) based on header row:
# CIS Control | CIS Safeguard | Asset Type | Security Function | Title | Description | IG1 | IG2 | IG3
COL_CONTROL   = 0
COL_SAFEGUARD = 1
COL_ASSET     = 2
COL_FUNCTION  = 3
COL_TITLE     = 4
COL_DESC      = 5
COL_IG1       = 6
COL_IG2       = 7
COL_IG3       = 8


def _cell_to_id(cell) -> str | None:
    """
    Reconstruct the true CIS safeguard string from a cell.

    Excel stores X.10 as float X.1 (same value as X.1).  The only way to
    distinguish them is via the cell's number_format:
      - 'General' or '0.0#' style -> use minimal decimal repr (e.g. "3.1")
      - '0.00' -> two decimal places are significant -> use "3.10"
    """
    val = cell.value
    if val is None:
        return None
    fmt = getattr(cell, 'number_format', 'General') or 'General'
    try:
        f = float(val)
    except (TypeError, ValueError):
        return str(val).strip() or None

    # Number format '0.00' means the spreadsheet intentionally shows 2 decimals
    if '0.00' in fmt:
        s = f"{f:.2f}"   # e.g. 3.10, 13.10
    else:
        # Minimal representation: drop trailing zeros after decimal
        s = f"{f:.10f}".rstrip('0').rstrip('.')
        # Avoid floating-point noise (e.g. 3.0999999... -> 3.1)
        try:
            # Round to 2 significant decimal places for safety
            s = str(round(f, 2))
            if s.endswith('.0'):
                s = s[:-2]
        except Exception:
            pass
    return s if s else None


def load_safeguards(xlsx_path: Path) -> list[dict]:
    """Parse the CIS v8 spreadsheet and return safeguard dicts."""
    # Must open without read_only to access cell.number_format
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        log.error("Sheet '%s' not found in %s", SHEET_NAME, xlsx_path)
        sys.exit(1)

    ws = wb[SHEET_NAME]

    safeguards: list[dict] = []
    header_skipped = False

    for row in ws.iter_rows(min_row=1, values_only=False):
        if not header_skipped:
            if row[COL_CONTROL].value == "CIS Control":
                header_skipped = True
            continue

        sfg_cell = row[COL_SAFEGUARD]
        if sfg_cell.value is None:
            continue  # control group header row

        sfg_id = _cell_to_id(sfg_cell)
        if not sfg_id or "." not in sfg_id:
            continue

        control_id = f"CIS {sfg_id}"

        def _x(cell_val) -> bool:
            return bool(cell_val and str(cell_val).strip().lower() == 'x')

        ig1 = _x(row[COL_IG1].value)
        ig2 = _x(row[COL_IG2].value)
        ig3 = _x(row[COL_IG3].value)

        if ig1:
            min_ig = 1
        elif ig2:
            min_ig = 2
        elif ig3:
            min_ig = 3
        else:
            min_ig = None

        safeguards.append({
            "control_id":  control_id,
            "title":       str(row[COL_TITLE].value or "").strip(),
            "description": str(row[COL_DESC].value or "").strip() or None,
            "domain":      str(row[COL_ASSET].value or "").strip() or None,
            "subdomain":   str(row[COL_FUNCTION].value or "").strip() or None,
            "min_ig":      min_ig,
        })

    log.info("Parsed %d safeguards from %s", len(safeguards), xlsx_path.name)
    return safeguards


def fw_id(con: sqlite3.Connection, short_name: str) -> str | None:
    row = con.execute(
        "SELECT id FROM compliance_frameworks WHERE short_name=?", (short_name,)
    ).fetchone()
    return row[0] if row else None


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        log.error("Spreadsheet not found: %s", xlsx_path)
        sys.exit(1)
    if not DB_PATH.exists():
        log.error("DB not found: %s", DB_PATH)
        sys.exit(1)

    safeguards = load_safeguards(xlsx_path)
    if not safeguards:
        log.error("No safeguards parsed — check spreadsheet format.")
        sys.exit(1)

    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA foreign_keys=OFF")

    try:
        # Fix cis8 kind: framework -> catalog
        cis8_id = fw_id(con, "cis8")
        if not cis8_id:
            log.error("cis8 framework not found — run app first.")
            sys.exit(1)
        con.execute(
            "UPDATE compliance_frameworks SET kind='catalog' WHERE short_name='cis8'"
        )
        log.info("cis8 kind -> 'catalog'")

        # Load existing catalog_controls for cis8
        existing: dict[str, str] = {
            row[0]: row[1]
            for row in con.execute(
                "SELECT control_id, id FROM catalog_controls WHERE framework_id=?",
                (cis8_id,)
            ).fetchall()
        }
        log.info("Found %d existing cis8 catalog_controls rows.", len(existing))

        # Update descriptions and metadata where currently NULL/empty
        desc_updated = 0
        not_found = []
        for sfg in safeguards:
            cid = sfg["control_id"]
            if cid not in existing:
                not_found.append(cid)
                continue
            cc_id = existing[cid]
            con.execute(
                """UPDATE catalog_controls
                   SET description=COALESCE(NULLIF(description,''),?),
                       domain=COALESCE(NULLIF(domain,''),?),
                       subdomain=COALESCE(NULLIF(subdomain,''),?)
                   WHERE id=?""",
                (sfg["description"], sfg["domain"], sfg["subdomain"], cc_id)
            )
            desc_updated += 1
        log.info("Updated metadata on %d catalog_controls rows.", desc_updated)
        if not_found:
            log.warning("  %d safeguards not matched in DB: %s", len(not_found), not_found[:10])

        # Seed IG baselines
        ig_ids = {
            1: fw_id(con, "cis_ig1"),
            2: fw_id(con, "cis_ig2"),
            3: fw_id(con, "cis_ig3"),
        }
        for lvl, bid in ig_ids.items():
            if not bid:
                log.warning("  cis_ig%d baseline not found — skipping.", lvl)

        counts = {1: 0, 2: 0, 3: 0}
        for sfg in safeguards:
            cid    = sfg["control_id"]
            min_ig = sfg["min_ig"]
            if min_ig is None or cid not in existing:
                continue
            cc_id = existing[cid]
            for ig_level in range(min_ig, 4):
                bl_id = ig_ids.get(ig_level)
                if not bl_id:
                    continue
                con.execute(
                    "INSERT OR IGNORE INTO baseline_controls "
                    "(baseline_id, catalog_control_id, is_required, created_at) "
                    "VALUES (?,?,1,CURRENT_TIMESTAMP)",
                    (bl_id, cc_id)
                )
                counts[ig_level] += 1

        for lvl in (1, 2, 3):
            log.info("  cis_ig%d  -> %d baseline_controls rows", lvl, counts[lvl])

        con.commit()

        # Summary
        log.info("\n=== CIS IG baseline_controls summary ===")
        for sn in ("cis_ig1", "cis_ig2", "cis_ig3"):
            bid = fw_id(con, sn)
            if bid:
                n = con.execute(
                    "SELECT COUNT(*) FROM baseline_controls WHERE baseline_id=?", (bid,)
                ).fetchone()[0]
                log.info("  %-12s  %d controls", sn, n)

    except Exception as e:
        con.rollback()
        log.error("Failed — rolled back: %s", e)
        raise
    finally:
        con.execute("PRAGMA foreign_keys=ON")
        con.close()


if __name__ == "__main__":
    main()
